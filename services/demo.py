import datetime
import re
import math

from openpyxl import load_workbook
from os.path import join, abspath
from lexicon.lexicon_ru import ABBREVIATION

class NotAllData(Exception):
    pass


# Функция высчитывает на каком курсе учится группа без учета високосного года
def get_course_number(name_group: str) -> int:
    year_admission = int(f'20{name_group[9:]}')
    start_admission = datetime.date(year_admission, 8, 1)
    study_days = (datetime.date.today()-start_admission).days
    course = math.ceil(study_days / 365)
    return course


def get_cells_group(worksheet):
    cells = [cell for cell in next(
        worksheet.iter_rows(min_row=5, min_col=5, max_row=5, max_col=worksheet.max_column))]
    return cells


# Возвращает словарь с курсами внутри которых отсортированы группы с указанием столбца
def get_courses_group(cells, index_group) -> dict[int, list[dict[str, int]]]:
    courses = {1: [], 2: [], 3: [], 4: [], 5: [], 6: [], 7: []}
    for cell in cells:
        group = cell.value
        if re.match(fr'{index_group}', group):
            course_group = get_course_number(group)
            courses[course_group].append(dict({group: cell.column}))
    return courses


# Функция по заданной колонке группы выдает расписание на неделю
def get_timetable(sheet, column_group: int) -> dict[str, list]:
    timetable_week = {'Понедельник': [], 'Вторник': [], 'Среда': [],
                      'Четверг': [], 'Пятница': [], 'Суббота': []}
    time_table = [cell.value for cell in next(
        sheet.iter_cols(min_row=6, min_col=column_group, max_row=sheet.max_row, max_col=column_group))]
    lessons_count = 0
    for timetable_day in timetable_week:
        for lesson in range(lessons_count, lessons_count+8):
            timetable_week[timetable_day].append(time_table[lesson])
        lessons_count += 8
    return timetable_week


# Функция возвращает все ячейки первой страницы
def get_first_sheet(file_name):
    data_path = abspath(join('.', file_name))
    wb = load_workbook(filename=data_path,  data_only=True)
    wsn = wb.sheetnames
    name_first_sheet = str(wsn[0])
    ws = wb[name_first_sheet]
    wb.close()
    return ws


def get_bottom_group_course(directions, course):
    file = 'demo_timetable.xlsx'
    ws_1 = get_first_sheet(file)
    cells_group = get_cells_group(ws_1)
    directions_groups = get_courses_group(cells_group, directions)
    group_dict = directions_groups[course]
    group_list = []
    for i in range(0, len(group_dict)):
        name_group, col = list(directions_groups[course][i].items())[0]
        group_list.append(name_group)
    return group_list


# print(get_bottom_group_course(ABBREVIATION['Строительство'], 3))

# group, col = list(arch_groups[3][0].items())[0]
#
# timetable_arch_group = get_timetable(ws, col)

#
# for i in timetable_arch_group:
#     print(i)
#     print('_____________')
#     for j in timetable_arch_group[i]:
#         print(j)


